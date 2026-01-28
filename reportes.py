# reportes.py
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from datetime import datetime
import streamlit as st
from constantes import COLUMNAS_EXCEL, LISTAS_OPCIONES

def generar_acta_excel(datos, df_completo):
    try:
        try:
            wb = openpyxl.load_workbook('Acta de Asignación Equipos - V3.xlsx')
        except:
            return None

        ws = wb.active
        ws['P7'] = str(datos.get('USUARIO', '')).upper()
        ws['G12'] = datetime.now().strftime('%d/%m/%Y')
        ws['T12'] = datos.get('UBICACIÓN','-')
        ws['AG12'] = datos.get('DIRECCIÓN','-')
        ws['G14'] = datos.get('ÁREA','-')
        ws['T14'] = datos.get('ACTA DE  ASIGNACIÓN','-')
        
        usuario_actual = datos.get('USUARIO')
        if usuario_actual and len(usuario_actual) > 3:
            e_u = df_completo[df_completo['USUARIO'] == usuario_actual]
            mons = e_u[e_u['TIPO'].str.contains("MONITOR", case=False, na=False)]['NRO DE SERIE'].tolist()
            ws['Q18'] = " / ".join(mons) if mons else datos.get('COMPONENTE', '-')
        else:
            ws['Q18'] = datos.get('COMPONENTE', '-')
        
        t_p = str(datos.get('TIPO', '')).upper()
        ws['J20'] = "X" if any(x in t_p for x in ["AIO", "ALL IN ONE"]) else ""
        ws['J21'] = "X" if any(x in t_p for x in ["DESKTOP", "CPU"]) else ""
        ws['J22'] = "X" if "LAPTOP" in t_p else ""
        
        ws['R20'] = datos.get('NUEVO ACTIVO','-')
        ws['R21'] = datos.get('NRO DE SERIE','-')
        ws['R22'] = datos.get('EQUIPO','-')

        acc = str(datos.get('ACCESORIOS', '')).lower() 
        if "LAPTOP" in t_p: ws['O24'] = "X"
        else: ws['O24'] = "X" if "cargador" in acc else ""
        
        ws['R24'] = "X" if "cadena" in acc or "candado" in acc else ""
        ws['U24'] = "X" if "mouse" in acc or "ratón" in acc else ""
        ws['X24'] = "X" if "mochila" in acc or "maletín" in acc else ""
        ws['Z24'] = "X" if "teclado" in acc else ""

        out = BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception as e:
        return None

def generar_plantilla_carga():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNAS_EXCEL)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    validaciones = {
        "TIPO": LISTAS_OPCIONES["TIPO"],
        "ESTADO": LISTAS_OPCIONES["ESTADO"],
        "MARCA": LISTAS_OPCIONES["MARCA"],
        "ÁREA": LISTAS_OPCIONES["ÁREA"]
    }
    
    for col_name, opciones in validaciones.items():
        if col_name in COLUMNAS_EXCEL:
            col_idx = COLUMNAS_EXCEL.index(col_name) + 1
            letra = openpyxl.utils.get_column_letter(col_idx)
            formula = f'"{",".join(opciones)}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letra}2:{letra}1000")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
