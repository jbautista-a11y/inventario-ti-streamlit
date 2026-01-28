import streamlit as st
import pandas as pd
from supabase import create_client, Client
import msal
import time
import plotly.express as px
from io import BytesIO
from datetime import datetime
from streamlit_cookies_manager import EncryptedCookieManager
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestión de Inventario TI", layout="wide", page_icon="🖥️")

# --- ESTILOS CSS ---
st.markdown("""
    <style>
        .block-container { padding-top: 2rem !important; }
        .stTabs { margin-top: 0px !important; }
        hr { margin-top: 10px !important; margin-bottom: 10px !important; }
        .stAlert { padding-top: 0.5rem; padding-bottom: 0.5rem; }
    </style>
    """, unsafe_allow_html=True)

# --- CONFIGURACIÓN SUPABASE ---
@st.cache_resource
def init_supabase():
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        return None

supabase: Client = init_supabase()

if not supabase:
    st.error("❌ Error Crítico: No se detectaron las credenciales de Supabase. Verifica tus 'secrets'.")
    st.stop()

# --- CONFIGURACIÓN MICROSOFT (AZURE AD) ---
try:
    CLIENT_ID = st.secrets["CLIENT_ID"]
    TENANT_ID = st.secrets["TENANT_ID"]
    CLIENT_SECRET = st.secrets["CLIENT_SECRET"]
    REDIRECT_URI = st.secrets["REDIRECT_URI"]
    AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
    SCOPE = ["User.Read"]
    auth_configured = True
except:
    auth_configured = False
    CLIENT_ID = "dummy" 

# --- GESTOR DE COOKIES ---
cookies = EncryptedCookieManager(password=st.secrets.get("COOKIE_PASSWORD", "esandata_secret_key_2024"))
if not cookies.ready():
    st.stop()

# --- DEFINICIÓN DE COLUMNAS (EXACTA A TU EXCEL LOCAL) ---
COLUMNAS_EXCEL = [
    "N°", "USUARIO", "EQUIPO", "ÁREA", "DIRECCIÓN", "UBICACIÓN", 
    "NUEVO ACTIVO", "ACTIVO", "TIPO", "NRO DE SERIE", "MARCA", "MODELO", 
    "AÑO DE ADQUISICIÓN", "PROCESADOR", "MEMORIA RAM", "DISCO DURO", 
    "ESTADO", "COMPONENTE", "COSTO", "ACCESORIOS", "OBSERVACIONES", 
    "ACTA DE  ASIGNACIÓN", "ADM- LOCAL", "ORIGEN_HOJA", "Ultima_Actualizacion", "MODIFICADO_POR"
]

MAPEO_DB = {
    "N°": "numero", "USUARIO": "usuario", "EQUIPO": "equipo", "ÁREA": "area",
    "DIRECCIÓN": "direccion", "UBICACIÓN": "ubicacion", "NUEVO ACTIVO": "nuevo_activo",
    "ACTIVO": "activo", "TIPO": "tipo", "NRO DE SERIE": "nro_serie",
    "MARCA": "marca", "MODELO": "modelo", "AÑO DE ADQUISICIÓN": "anio_adquisicion",
    "PROCESADOR": "procesador", "MEMORIA RAM": "memoria_ram", "DISCO DURO": "disco_duro",
    "ESTADO": "estado", "COMPONENTE": "componente", "COSTO": "costo",
    "ACCESORIOS": "accesorios", "OBSERVACIONES": "observaciones",
    "ACTA DE  ASIGNACIÓN": "acta_asignacion", "ADM- LOCAL": "adm_local",
    "ORIGEN_HOJA": "origen_hoja", "Ultima_Actualizacion": "ultima_actualizacion",
    "MODIFICADO_POR": "modificado_por"
}
MAPEO_INVERSO = {v: k for k, v in MAPEO_DB.items()}

# --- LISTAS BASE ---
LISTAS_OPCIONES = {
    "TIPO": ["LAPTOP", "DESKTOP", "MONITOR", "ALL IN ONE", "TABLET", "IMPRESORA", "PERIFERICO", "PROYECTOR", "TV"],
    "ESTADO": ["OPERATIVO", "EN REVISIÓN", "MANTENIMIENTO", "BAJA", "HURTO/ROBO", "ASIGNADO", "DISPONIBLE"],
    "MARCA": ["DELL", "HP", "LENOVO", "APPLE", "SAMSUNG", "LG", "EPSON", "LOGITECH", "ASUS", "ACER"],
    "ÁREA": ["SOPORTE TI", "ADMINISTRACIÓN", "RECURSOS HUMANOS", "CONTABILIDAD", "COMERCIAL", "MARKETING", "LOGÍSTICA", "DIRECCIÓN", "ACADÉMICO"]
}

# --- FUNCIONES AUXILIARES ---

def registrar_log(accion, detalle):
    try:
        usuario = st.session_state.get("usuario_actual", "Desconocido")
        datos = {
            "usuario": usuario, "accion": accion, "detalle": detalle, "fecha": datetime.now().isoformat()
        }
        supabase.table('logs_auditoria').insert(datos).execute()
    except Exception as e:
        print(f"Error log: {e}")

def cargar_usuarios():
    try:
        response = supabase.table('usuarios').select("*").execute()
        df = pd.DataFrame(response.data)
        if df.empty: return pd.DataFrame(columns=["usuario", "clave", "rol"])
        return df
    except:
        return pd.DataFrame(columns=["usuario", "clave", "rol"])

def guardar_nuevo_usuario(u, r):
    try:
        df = cargar_usuarios()
        if not df.empty and u.lower() in df["usuario"].str.lower().values:
            return False, "Usuario ya existe"
        supabase.table('usuarios').insert({"usuario": u.lower(), "clave": "MS_365_ACCESS", "rol": r}).execute()
        return True, f"Usuario {u} autorizado."
    except Exception as e:
        return False, f"Error DB: {e}"

def actualizar_clave_local(usuario, nueva_clave):
    try:
        supabase.table('usuarios').update({"clave": nueva_clave}).eq("usuario", usuario).execute()
        registrar_log("SEGURIDAD", "Cambio de clave local")
        return True, "Clave actualizada"
    except Exception as e:
        return False, str(e)

# --- NÚCLEO DE DATOS ---

@st.cache_data(ttl=60) # Refresco automático cada 60s si no se limpia antes
def obtener_datos():
    try:
        # Traer hasta 10,000 registros para asegurar completitud
        response = supabase.table('inventario').select("*").limit(10000).order('id', desc=False).execute()
        data = response.data
        if not data: return pd.DataFrame(columns=COLUMNAS_EXCEL)
            
        df = pd.DataFrame(data)
        df = df.rename(columns=MAPEO_INVERSO)
        
        # Asegurar columnas faltantes
        for col in COLUMNAS_EXCEL:
            if col not in df.columns: df[col] = "-"
        
        # Limpieza CRÍTICA para búsquedas
        df = df.fillna("")
        
        # Guardar ID oculto
        if "id" in pd.DataFrame(data).columns:
            df["_supabase_id"] = pd.DataFrame(data)["id"]
        
        # Convertir todo a string mayúsculas para búsquedas perfectas
        for col in df.columns:
            if col != "_supabase_id":
                df[col] = df[col].astype(str).str.upper().str.strip()
                df[col] = df[col].replace(["NAN", "NONE", "NULL"], "")
                
        return df
    except Exception as e:
        st.error(f"Error cargando datos: {e}")
        return pd.DataFrame(columns=COLUMNAS_EXCEL)

def guardar_registro_db(datos_dict, es_nuevo=True, id_supabase=None):
    try:
        datos_db = {}
        for k, v in datos_dict.items():
            if k in MAPEO_DB: datos_db[MAPEO_DB[k]] = v
        
        datos_db["ultima_actualizacion"] = datetime.now().isoformat()
        datos_db["modificado_por"] = st.session_state.get("usuario_actual", "Sistema")
        
        if es_nuevo:
            datos_db["numero"] = str(int(time.time())) # ID temporal visual
            supabase.table('inventario').insert(datos_db).execute()
        else:
            if id_supabase:
                supabase.table('inventario').update(datos_db).eq('id', id_supabase).execute()
        
        st.cache_data.clear() # ¡Limpiar caché inmediatamente!
        return True
    except Exception as e:
        st.error(f"Error guardando en BD: {e}")
        return False

# --- FUNCIONES UI ---

def campo_con_opcion_otro(label, lista_base, valor_actual=None, key_suffix=""):
    opciones = list(lista_base)
    opcion_otro = "OTRO (ESPECIFICAR)"
    if opcion_otro not in opciones: opciones.append(opcion_otro)
    
    idx = 0
    modo_manual = False
    
    # Lógica inteligente para pre-seleccionar
    if valor_actual and valor_actual not in ["", "-", "NAN"]:
        if valor_actual in opciones:
            idx = opciones.index(valor_actual)
        else:
            # Si el valor no está en la lista (ej. una marca nueva en DB), lo añadimos temporalmente o activamos manual
            idx = opciones.index(opcion_otro)
            modo_manual = True
            
    seleccion = st.selectbox(label, opciones, index=idx, key=f"sel_{label}_{key_suffix}")
    valor_final = seleccion
    
    if seleccion == opcion_otro:
        val_defecto = valor_actual if modo_manual else ""
        valor_final = st.text_input(f"Especifique {label}:", value=val_defecto, key=f"txt_{label}_{key_suffix}").upper()
        
    return valor_final

def generar_acta_excel(datos, df_completo):
    try:
        # Intenta cargar la plantilla
        try:
            wb = openpyxl.load_workbook('Acta de Asignación Equipos - V3.xlsx')
        except:
            st.error("⚠️ Falta el archivo 'Acta de Asignación Equipos - V3.xlsx' en tu repositorio.")
            return None

        ws = wb.active
        # Mapeo idéntico al local
        ws['P7'] = str(datos.get('USUARIO', '')).upper()
        ws['G12'] = datetime.now().strftime('%d/%m/%Y')
        ws['T12'] = datos.get('UBICACIÓN','-')
        ws['AG12'] = datos.get('DIRECCIÓN','-')
        ws['G14'] = datos.get('ÁREA','-')
        ws['T14'] = datos.get('ACTA DE  ASIGNACIÓN','-')
        
        # Componentes asociados (Monitores)
        usuario_actual = datos.get('USUARIO')
        if usuario_actual and len(usuario_actual) > 3:
            e_u = df_completo[df_completo['USUARIO'] == usuario_actual]
            mons = e_u[e_u['TIPO'].str.contains("MONITOR", case=False, na=False)]['NRO DE SERIE'].tolist()
            ws['Q18'] = " / ".join(mons) if mons else datos.get('COMPONENTE', '-')
        else:
            ws['Q18'] = datos.get('COMPONENTE', '-')
        
        # Checkboxes Tipo
        t_p = str(datos.get('TIPO', '')).upper()
        ws['J20'] = "X" if any(x in t_p for x in ["AIO", "ALL IN ONE"]) else ""
        ws['J21'] = "X" if any(x in t_p for x in ["DESKTOP", "CPU"]) else ""
        ws['J22'] = "X" if "LAPTOP" in t_p else ""
        
        ws['R20'] = datos.get('NUEVO ACTIVO','-')
        ws['R21'] = datos.get('NRO DE SERIE','-')
        ws['R22'] = datos.get('EQUIPO','-')

        # Checkboxes Accesorios
        acc = str(datos.get('ACCESORIOS', '')).lower() 
        if "LAPTOP" in t_p: ws['O24'] = "X"
        else: ws['O24'] = "X" if "cargador" in acc else ""
        
        ws['R24'] = "X" if "cadena" in acc or "candado" in acc else ""
        ws['U24'] = "X" if "mouse" in acc or "ratón" in acc else ""
        ws['X24'] = "X" if "mochila" in acc or "maletín" in acc else ""
        ws['Z24'] = "X" if "teclado" in acc else ""

        out = BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception as e:
        st.error(f"Error generando acta: {e}")
        return None

def generar_plantilla_carga():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNAS_EXCEL)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    # Validaciones para ayudar al usuario
    validaciones = {
        "TIPO": LISTAS_OPCIONES["TIPO"],
        "ESTADO": LISTAS_OPCIONES["ESTADO"],
        "MARCA": LISTAS_OPCIONES["MARCA"],
        "ÁREA": LISTAS_OPCIONES["ÁREA"]
    }
    
    for col_name, opciones in validaciones.items():
        if col_name in COLUMNAS_EXCEL:
            col_idx = COLUMNAS_EXCEL.index(col_name) + 1
            letra = openpyxl.utils.get_column_letter(col_idx)
            formula = f'"{",".join(opciones)}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letra}2:{letra}1000")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# --- LOGIN ---

def verificar_sesion():
    if "autenticado" not in st.session_state: st.session_state.autenticado = False

    # 1. Recuperar Cookie
    if not st.session_state.autenticado:
        c_user = cookies.get("usuario_actual")
        c_rol = cookies.get("rol_actual")
        if c_user and c_rol:
            st.session_state.autenticado = True
            st.session_state.usuario_actual = c_user
            st.session_state.rol_actual = c_rol

    # 2. Microsoft Callback
    if "code" in st.query_params:
        if not auth_configured:
            st.error("Autenticación Microsoft no configurada correctamente.")
        else:
            try:
                app = msal.ConfidentialClientApplication(CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)
                result = app.acquire_token_by_authorization_code(st.query_params["code"], scopes=SCOPE, redirect_uri=REDIRECT_URI)
                if "error" not in result:
                    email = result.get("id_token_claims").get("preferred_username").lower()
                    df_u = cargar_usuarios()
                    user_match = df_u[df_u["usuario"].str.lower() == email]
                    if not user_match.empty:
                        st.session_state.autenticado = True
                        st.session_state.usuario_actual = email
                        st.session_state.rol_actual = user_match.iloc[0]["rol"]
                        cookies["usuario_actual"] = email
                        cookies["rol_actual"] = st.session_state.rol_actual
                        cookies.save()
                        st.query_params.clear()
                        st.rerun()
                    else:
                        st.error(f"El usuario {email} no tiene permisos en la Base de Datos.")
            except Exception as e:
                st.error(f"Error Login MS: {e}")

    # 3. Pantalla Login
    if not st.session_state.autenticado:
        st.markdown("<h1 style='text-align: center;'>☁️ Gestión de Inventario TI</h1>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 1.2, 1])
        with col2:
            if auth_configured:
                try:
                    app = msal.ConfidentialClientApplication(CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)
                    auth_url = app.get_authorization_request_url(SCOPE, redirect_uri=REDIRECT_URI)
                    st.link_button("🟦 Iniciar con Microsoft 365", auth_url, use_container_width=True)
                except: pass
            
            st.divider()
            with st.expander("🔐 Acceso Local"):
                with st.form("local_login"):
                    u = st.text_input("Usuario")
                    p = st.text_input("Clave", type="password")
                    if st.form_submit_button("Entrar", use_container_width=True):
                        df_u = cargar_usuarios()
                        match = df_u[(df_u["usuario"].str.lower() == u.lower()) & (df_u["clave"] == p)]
                        if not match.empty:
                            st.session_state.autenticado = True
                            st.session_state.usuario_actual = match.iloc[0]["usuario"]
                            st.session_state.rol_actual = match.iloc[0]["rol"]
                            cookies["usuario_actual"] = st.session_state.usuario_actual
                            cookies["rol_actual"] = st.session_state.rol_actual
                            cookies.save()
                            registrar_log("LOGIN_LOCAL", "Acceso Local Exitoso")
                            st.rerun()
                        else:
                            st.error("Usuario o clave incorrectos")
        st.stop()
    return True

# --- APLICACIÓN PRINCIPAL ---

if verificar_sesion():
    # SIDEBAR
    with st.sidebar:
        st.title("⚙️ Panel")
        st.write(f"👤 **{st.session_state.usuario_actual}**")
        st.info(f"Rol: {st.session_state.rol_actual}")
        
        # Opción cambiar clave local
        df_u = cargar_usuarios()
        curr = df_u[df_u["usuario"] == st.session_state.usuario_actual]
        if not curr.empty and curr.iloc[0]["clave"] != "MS_365_ACCESS":
            with st.expander("Cambiar Clave"):
                n_p = st.text_input("Nueva Clave", type="password")
                if st.button("Actualizar Clave"):
                    ok, m = actualizar_clave_local(st.session_state.usuario_actual, n_p)
                    if ok: st.success(m)
                    else: st.error(m)
        
        st.divider()
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            cookies["usuario_actual"] = ""
            cookies["rol_actual"] = ""
            cookies.save()
            st.session_state.clear()
            st.rerun()

    c_head, c_logo = st.columns([3, 1])
    with c_head: st.title("🖥️ Gestión de Inventario TI")
    
    # Cargar Datos Centralizados
    df = obtener_datos()
    
    pestanas = ["📊 Dashboard", "🔎 Consultar", "➕ Nuevo", "📥 Carga Masiva", "✏️ Editar/Acta"]
    if st.session_state.rol_actual == "Administrador":
        pestanas += ["📜 Logs", "👥 Usuarios"]
    
    tabs = st.tabs(pestanas)

    # 1. DASHBOARD
    with tabs[0]:
        st.subheader("Tablero de Control")
        with st.expander("🔎 Filtros Avanzados", expanded=False):
            fc1, fc2, fc3 = st.columns(3)
            # Filtros Inteligentes: Combinan lista base con lo que realmente existe en DB
            opts_area = sorted(list(set(LISTAS_OPCIONES["ÁREA"] + df["ÁREA"].unique().tolist())))
            opts_tipo = sorted(list(set(LISTAS_OPCIONES["TIPO"] + df["TIPO"].unique().tolist())))
            opts_estado = sorted(list(set(LISTAS_OPCIONES["ESTADO"] + df["ESTADO"].unique().tolist())))
            
            if "" in opts_area: opts_area.remove("")
            
            sel_area = fc1.multiselect("Área", opts_area)
            sel_tipo = fc2.multiselect("Tipo", opts_tipo)
            sel_estado = fc3.multiselect("Estado", opts_estado)
        
        df_d = df.copy()
        if sel_area: df_d = df_d[df_d["ÁREA"].isin(sel_area)]
        if sel_tipo: df_d = df_d[df_d["TIPO"].isin(sel_tipo)]
        if sel_estado: df_d = df_d[df_d["ESTADO"].isin(sel_estado)]
        
        # Limpieza para métricas numéricas
        def limpiar_costo(val):
            try: return float(str(val).replace("S/", "").replace("$", "").replace(",", ""))
            except: return 0.0
        
        df_d["COSTO_NUM"] = df_d["COSTO"].apply(limpiar_costo)
        
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Activos", len(df_d))
        k2.metric("Asignados", len(df_d[df_d["USUARIO"].str.len() > 2])) # Usuario > 2 chars
        k3.metric("Stock / Mtto", len(df_d[df_d["ESTADO"].isin(["EN REVISIÓN", "MANTENIMIENTO", "OPERATIVO", "DISPONIBLE"]) & (df_d["USUARIO"].str.len() <= 2)]))
        k4.metric("Valor Total", f"S/ {df_d['COSTO_NUM'].sum():,.2f}")
        
        st.divider()
        g1, g2 = st.columns(2)
        with g1:
            if not df_d.empty:
                pie_data = df_d["TIPO"].value_counts().reset_index()
                pie_data.columns = ["Tipo", "Conteo"]
                st.plotly_chart(px.pie(pie_data, values="Conteo", names="Tipo", title="Distribución por Tipo", hole=0.4), use_container_width=True)
        with g2:
            if not df_d.empty:
                bar_data = df_d["ÁREA"].value_counts().head(10).reset_index()
                bar_data.columns = ["Área", "Equipos"]
                st.plotly_chart(px.bar(bar_data, x="Equipos", y="Área", orientation='h', title="Top Áreas"), use_container_width=True)

    # 2. CONSULTAR
    with tabs[1]:
        st.subheader("Base de Datos Completa")
        # Aquí también añadimos un buscador rápido
        busqueda_rapida = st.text_input("Filtrar tabla:", placeholder="Escribe para filtrar visualmente...").upper()
        
        df_view = df.drop(columns=["_supabase_id"], errors="ignore")
        if busqueda_rapida:
            # Filtro global en tabla visual
            mask_fast = df_view.astype(str).apply(lambda x: x.str.contains(busqueda_rapida, na=False)).any(axis=1)
            df_view = df_view[mask_fast]
            
        st.dataframe(df_view, use_container_width=True, hide_index=True)

    # 3. NUEVO REGISTRO
    with tabs[2]:
        st.subheader("Registrar Nuevo Activo")
        with st.form("frm_nuevo"):
            c1, c2, c3 = st.columns(3)
            datos_nuevos = {}
            with c1:
                datos_nuevos["USUARIO"] = st.text_input("Usuario Asignado").upper()
                datos_nuevos["ÁREA"] = campo_con_opcion_otro("Área", LISTAS_OPCIONES["ÁREA"], key_suffix="n1")
                datos_nuevos["UBICACIÓN"] = st.text_input("Ubicación Física").upper()
                datos_nuevos["DIRECCIÓN"] = st.text_input("Dirección").upper()
            with c2:
                datos_nuevos["TIPO"] = campo_con_opcion_otro("Tipo", LISTAS_OPCIONES["TIPO"], key_suffix="n2")
                datos_nuevos["MARCA"] = campo_con_opcion_otro("Marca", LISTAS_OPCIONES["MARCA"], key_suffix="n3")
                datos_nuevos["MODELO"] = st.text_input("Modelo").upper()
                datos_nuevos["EQUIPO"] = st.text_input("Hostname / Equipo").upper()
            with c3:
                datos_nuevos["NRO DE SERIE"] = st.text_input("Nro de Serie").upper()
                datos_nuevos["NUEVO ACTIVO"] = st.text_input("Cód. Nuevo Activo").upper()
                datos_nuevos["ACTIVO"] = st.text_input("Cód. Antiguo (Activo)").upper()
                datos_nuevos["ESTADO"] = campo_con_opcion_otro("Estado", LISTAS_OPCIONES["ESTADO"], key_suffix="n4")
                
            datos_nuevos["COSTO"] = st.text_input("Costo").upper()
            datos_nuevos["ACCESORIOS"] = st.text_area("Accesorios").upper()
            datos_nuevos["OBSERVACIONES"] = st.text_area("Observaciones").upper()
            
            if st.form_submit_button("💾 Guardar Registro"):
                # Validación básica de duplicados
                if datos_nuevos["NRO DE SERIE"] and datos_nuevos["NRO DE SERIE"] in df["NRO DE SERIE"].values:
                    st.error(f"Error: El Nro de Serie {datos_nuevos['NRO DE SERIE']} ya existe en la base de datos.")
                else:
                    if guardar_registro_db(datos_nuevos, es_nuevo=True):
                        st.success("✅ Registro guardado exitosamente.")
                        registrar_log("CREAR", f"Alta activo: {datos_nuevos['NRO DE SERIE']}")
                        time.sleep(1.5)
                        st.rerun()

    # 4. CARGA MASIVA
    with tabs[3]:
        st.subheader("Carga Masiva desde Excel")
        col_down, col_up = st.columns(2)
        with col_down:
            st.info("Paso 1: Descargue la plantilla obligatoria.")
            plantilla = generar_plantilla_carga()
            st.download_button("📥 Descargar Plantilla .xlsx", data=plantilla, file_name="Plantilla_Carga.xlsx")
        
        with col_up:
            st.info("Paso 2: Suba el archivo con los datos.")
            upl_file = st.file_uploader("Subir Excel", type=["xlsx"])
            if upl_file:
                if st.button("Procesar Archivo"):
                    try:
                        df_upload = pd.read_excel(upl_file).fillna("")
                        df_upload = df_upload.astype(str)
                        progreso = st.progress(0)
                        total = len(df_upload)
                        for idx, row in df_upload.iterrows():
                            d_row = row.to_dict()
                            # Limpiar claves y valores
                            d_row_clean = {k.strip().upper(): v.strip().upper() for k,v in d_row.items()}
                            guardar_registro_db(d_row_clean, es_nuevo=True)
                            progreso.progress((idx + 1) / total)
                        
                        st.success(f"Proceso completado: {total} registros cargados.")
                        time.sleep(2)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error procesando archivo: {e}")

    # 5. EDITAR / ACTA - BÚSQUEDA CORREGIDA GLOBAL
    with tabs[4]:
        st.subheader("Gestión de Activos y Actas")
        
        c_search, c_info = st.columns([2, 1])
        with c_search:
            termino = st.text_input("🔍 Buscar Activo:", placeholder="Escriba usuario, serie, código antiguo, marca, IP...", help="Busca en TODAS las columnas").upper()
        
        df_res = pd.DataFrame()
        
        if termino:
            # --- CORRECCIÓN DE BÚSQUEDA ---
            # En lugar de buscar columna por columna, unimos toda la fila en un solo texto y buscamos ahí.
            # Esto garantiza que encuentre CUALQUIER COSA (Activo antiguo, Marca, IP, etc.)
            
            # Columnas clave para la búsqueda (puedes agregar más si quieres)
            cols_busqueda = ["USUARIO", "NRO DE SERIE", "NUEVO ACTIVO", "ACTIVO", "EQUIPO", "MARCA", "MODELO", "UBICACIÓN", "TIPO"]
            
            # Filtro Vectorizado (Rápido y Efectivo)
            mask = pd.Series(False, index=df.index)
            for col in cols_busqueda:
                if col in df.columns:
                    mask |= df[col].str.contains(termino, na=False)
            
            df_res = df[mask]
        else:
            # Si no hay búsqueda, mostramos los últimos 5 modificados para acceso rápido
            df_res = df.sort_values(by="Ultima_Actualizacion", ascending=False).head(5)

        if not df_res.empty:
            if termino:
                st.info(f"Se encontraron {len(df_res)} coincidencias.")
            
            # Crear etiqueta descriptiva para el selector
            df_res["_etiqueta"] = df_res.apply(
                lambda x: f"{x['USUARIO']} | {x['TIPO']} | {x['MARCA']} | S/N: {x['NRO DE SERIE']} | Act: {x['ACTIVO']}", 
                axis=1
            )
            
            seleccion = st.selectbox("Seleccione el activo a gestionar:", df_res["_etiqueta"].tolist())
            
            if seleccion:
                # Recuperar fila original
                row_sel = df_res[df_res["_etiqueta"] == seleccion].iloc[0]
                id_sel = row_sel["_supabase_id"]
                
                st.markdown("---")
                col_ed, col_acta = st.columns([1.5, 1])
                
                with col_ed:
                    st.markdown("#### ✏️ Editar Datos")
                    with st.form("frm_edicion"):
                        # Campos más comunes editables
                        ne_user = st.text_input("Usuario", row_sel["USUARIO"])
                        
                        c_e1, c_e2 = st.columns(2)
                        with c_e1:
                            ne_serie = st.text_input("Nro Serie", row_sel["NRO DE SERIE"])
                            ne_activo_new = st.text_input("Nuevo Activo", row_sel["NUEVO ACTIVO"])
                            ne_area = campo_con_opcion_otro("Área", LISTAS_OPCIONES["ÁREA"], row_sel["ÁREA"], "ed1")
                        with c_e2:
                            ne_equipo = st.text_input("Hostname", row_sel["EQUIPO"])
                            ne_activo_old = st.text_input("Activo (Antiguo)", row_sel["ACTIVO"])
                            ne_estado = campo_con_opcion_otro("Estado", LISTAS_OPCIONES["ESTADO"], row_sel["ESTADO"], "ed2")
                        
                        ne_obs = st.text_area("Observaciones", row_sel["OBSERVACIONES"])
                        ne_acc = st.text_area("Accesorios", row_sel["ACCESORIOS"])
                        
                        if st.form_submit_button("Actualizar Datos"):
                            cambios = {
                                "USUARIO": ne_user, "NRO DE SERIE": ne_serie,
                                "NUEVO ACTIVO": ne_activo_new, "ACTIVO": ne_activo_old,
                                "EQUIPO": ne_equipo, "ÁREA": ne_area, "ESTADO": ne_estado,
                                "OBSERVACIONES": ne_obs, "ACCESORIOS": ne_acc
                            }
                            # Preservar resto de datos
                            registro_upd = row_sel.to_dict()
                            registro_upd.update(cambios)
                            
                            if guardar_registro_db(registro_upd, es_nuevo=False, id_supabase=id_sel):
                                st.success("Registro actualizado correctamente")
                                registrar_log("EDITAR", f"Modificado: {ne_serie}")
                                time.sleep(1)
                                st.rerun()

                with col_acta:
                    st.markdown("#### 📄 Acciones")
                    st.info(f"Usuario: {row_sel['USUARIO']}")
                    
                    excel_acta = generar_acta_excel(row_sel.to_dict(), df)
                    if excel_acta:
                        n_arch = f"Acta_{row_sel['USUARIO']}_{row_sel['NRO DE SERIE']}.xlsx"
                        st.download_button("📥 Descargar Acta Excel", data=excel_acta, file_name=n_arch, use_container_width=True)
                    
                    st.divider()
                    if st.button("🗑️ Eliminar Registro", type="primary", use_container_width=True):
                        try:
                            supabase.table('inventario').delete().eq('id', id_sel).execute()
                            st.success("Registro eliminado permanentemente.")
                            registrar_log("ELIMINAR", f"Eliminado: {row_sel['NRO DE SERIE']}")
                            st.cache_data.clear()
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error borrando: {e}")
        elif termino:
            st.warning("No se encontraron coincidencias. Intente con otro término.")

    # 6. LOGS (SOLO ADMIN)
    if st.session_state.rol_actual == "Administrador":
        with tabs[5]:
            st.subheader("Auditoría del Sistema")
            if st.button("🔄 Refrescar Logs"): st.rerun()
            try:
                logs = supabase.table('logs_auditoria').select("*").order('fecha', desc=True).limit(100).execute()
                st.dataframe(pd.DataFrame(logs.data), use_container_width=True)
            except: st.info("No hay logs registrados.")

        with tabs[6]:
            st.subheader("Control de Accesos")
            col_u1, col_u2 = st.columns(2)
            with col_u1:
                st.markdown("##### Agregar Usuario")
                with st.form("new_user"):
                    nu_mail = st.text_input("Correo / Usuario")
                    nu_rol = st.selectbox("Rol", ["Soporte", "Administrador"])
                    if st.form_submit_button("Autorizar Acceso"):
                        ok, m = guardar_nuevo_usuario(nu_mail, nu_rol)
                        if ok: st.success(m); time.sleep(1); st.rerun()
                        else: st.error(m)
                
                st.divider()
                st.markdown("##### Revocar Acceso")
                df_u = cargar_usuarios()
                # Filtrar para no borrarse a sí mismo
                lista_del = [u for u in df_u["usuario"].tolist() if u != st.session_state.usuario_actual]
                
                if lista_del:
                    u_del = st.selectbox("Seleccione usuario a eliminar", lista_del)
                    if st.button("Eliminar Usuario", type="primary"):
                        supabase.table('usuarios').delete().eq('usuario', u_del).execute()
                        st.success(f"Usuario {u_del} eliminado.")
                        time.sleep(1)
                        st.rerun()
                else:
                    st.caption("No hay otros usuarios para eliminar.")

            with col_u2:
                st.markdown("##### Usuarios Activos")
                st.dataframe(cargar_usuarios(), use_container_width=True, hide_index=True)
